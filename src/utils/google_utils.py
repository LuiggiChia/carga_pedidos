import os
import json
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from googleapiclient.http import MediaIoBaseUpload
from google.oauth2.service_account import Credentials

SCOPES = ["https://www.googleapis.com/auth/drive"]


def load_google_credentials(base_dir: str, logger):
    logger.info("Cargando credenciales de Google")

    credentials_path = os.path.join(base_dir, "config/credentials.json")

    with open(credentials_path, "r") as file:
        credentials_data = json.load(file)

    logger.info("Credenciales cargadas correctamente")

    return Credentials.from_service_account_info(credentials_data, scopes=SCOPES)


def create_drive_service(credentials, logger):
    logger.info("Creando servicio de Google Drive")

    service = build("drive", "v3", credentials=credentials)

    logger.info("Servicio de Google Drive creado")

    return service


def upload_file_to_drive(service, base_dir: str, logger):

    credentials_folder_drive = os.path.join(
        base_dir, "config", "credentials_folder_drive.json"
    )

    with open(credentials_folder_drive, "r", encoding="utf-8") as file:
        folder_drive_credentials = json.load(file)

    folder_id = folder_drive_credentials["folder_id_raw"]

    raw_path = os.path.join(base_dir, "data", "raw")

    csv_file = next(f for f in os.listdir(raw_path) if f.lower().endswith(".csv"))

    file_path = os.path.join(raw_path, csv_file)

    media = MediaFileUpload(file_path, mimetype="text/csv")

    file = (
        service.files()
        .create(
            body={"name": csv_file, "parents": [folder_id]},
            media_body=media,
            fields="id,name",
            supportsAllDrives=True,
        )
        .execute()
    )

    logger.info(f"Subido: {file['name']}")
    logger.info(f"ID: {file['id']}")

    return file["id"]


def get_plantilla_carga_pedidos_and_upload_to_drive(
    service,
    base_dir: str,
    df: pd.DataFrame,
    logger,
    day_of_report: datetime = datetime.now(),
):

    report_date = day_of_report.strftime("%Y%m%d")

    credentials_folder_drive = os.path.join(
        base_dir,
        "config",
        "credentials_folder_drive.json",
    )

    with open(credentials_folder_drive, "r", encoding="utf-8") as file:
        folder_drive_credentials = json.load(file)

    folder_id_utils = folder_drive_credentials["folder_id_utils"]
    folder_id_factoring = folder_drive_credentials["folder_id_factoring"]

    query = (
        f"name = 'Plantilla Carga Pedidos.xlsx' "
        f"and '{folder_id_utils}' in parents "
        f"and trashed = false"
    )

    results = (
        service.files()
        .list(
            q=query,
            fields="files(id)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        .execute()
    )

    files = results.get("files", [])

    if not files:
        logger.error("File 'Plantilla Carga Pedidos.xlsx' was not found.")
        return None

    file_id = files[0]["id"]

    # Descargar plantilla
    file_bytes = service.files().get_media(fileId=file_id).execute()

    # Abrir plantilla
    excel_buffer = BytesIO(file_bytes)

    wb = load_workbook(excel_buffer)

    ws = wb["Carga Pedidos"]

    # Escribir datos desde fila 2
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Guardar workbook modificado en memoria
    output = BytesIO()

    wb.save(output)

    output.seek(0)

    # Subir archivo a Drive
    media = MediaIoBaseUpload(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    uploaded_file = (
        service.files()
        .create(
            body={
                "name": f"CargaPedidos_{report_date}.xlsx",
                "parents": [folder_id_factoring],
            },
            media_body=media,
            supportsAllDrives=True,
            fields="id,name",
        )
        .execute()
    )

    logger.info(
        f"Se subió correctamente '{uploaded_file['name']}' "
        f"(ID: {uploaded_file['id']})"
    )

    return uploaded_file["id"]
