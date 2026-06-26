import os
import json
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from googleapiclient.http import MediaIoBaseUpload


def upload_dataframe_to_template_drive(
    service,
    # base_dir: str,
    df: pd.DataFrame,
    folder_id_utils: str,
    folder_output_id: str,
    excel_input_name: str,
    sheet_name: str,
    excel_output_name: str,
    logger,
    day_of_report: datetime = datetime.now(),
):

    if isinstance(day_of_report, str):
        day_of_report = datetime.strptime(day_of_report, "%d/%m/%Y")

    report_date = day_of_report.strftime("%Y%m%d")

    # credentials_folder_drive = os.path.join(
    #     base_dir,
    #     "config",
    #     "credentials_folder_drive.json",
    # )

    # with open(credentials_folder_drive, "r", encoding="utf-8") as file:
    #     folder_drive_credentials = json.load(file)

    # folder_id_utils = folder_drive_credentials["folder_id_utils"]
    # folder_id_factoring = folder_drive_credentials["folder_id_factoring"]

    query = (
        f"name = '{excel_input_name}' "
        f"and '{folder_id_utils}' in parents "
        f"and trashed = false"
    )

    results = (
        service.files()
        .list(
            q=query,
            fields="files(id)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        .execute()
    )

    files = results.get("files", [])

    if not files:
        logger.error(f"File '{excel_input_name}' was not found.")
        return None

    file_id = files[0]["id"]

    # Descargar plantilla
    file_bytes = service.files().get_media(fileId=file_id).execute()

    # Abrir plantilla
    excel_buffer = BytesIO(file_bytes)

    wb = load_workbook(excel_buffer)

    # ws = wb["Carga Pedidos"]
    ws = wb[sheet_name]

    # Escribir datos desde fila 2
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Guardar workbook modificado en memoria
    output = BytesIO()

    wb.save(output)

    output.seek(0)

    # Subir archivo a Drive
    media = MediaIoBaseUpload(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    uploaded_file = (
        service.files()
        .create(
            body={
                "name": f"{excel_output_name}_{report_date}.xlsx",
                "parents": [folder_output_id],
            },
            media_body=media,
            supportsAllDrives=True,
            fields="id,name",
        )
        .execute()
    )

    logger.info(
        f"Se subió correctamente '{uploaded_file['name']}' "
        f"(ID: {uploaded_file['id']})"
    )

    return uploaded_file["id"]
